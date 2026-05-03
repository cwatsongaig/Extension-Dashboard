"""
Extract data from Excel reports and generate data.js for BondBox prototype.
Reads 6 Excel files from ./Data/ and outputs a JavaScript file with all data arrays.
"""

import openpyxl
import json
import re
from collections import defaultdict
from datetime import datetime

DATA_DIR = r'C:\Development\New Extension Build\Data'

# 8 selected user profiles
USER_PROFILES = [
    {'username': 'dhellmann', 'fullName': 'Doug Hellmann', 'branch': 'KENTUCKY', 'role': 'Underwriter', 'avatar': 'DH'},
    {'username': 'jliggett', 'fullName': 'Jake Liggett', 'branch': 'CINCINNATI', 'role': 'Underwriter', 'avatar': 'JL'},
    {'username': 'wwilson', 'fullName': 'Wade Wilson', 'branch': 'ORANGE', 'role': 'Sr. Underwriter', 'avatar': 'WW'},
    {'username': 'kwittler', 'fullName': 'Kathy Wittler', 'branch': 'ORANGE', 'role': 'Underwriter', 'avatar': 'KW'},
    {'username': 'jwhipkey', 'fullName': 'Jonathon Whipkey', 'branch': 'TENNESSEE', 'role': 'Underwriter', 'avatar': 'JW'},
    {'username': 'akveton', 'fullName': 'Adam Kveton', 'branch': 'CHICAGO CONTRACT', 'role': 'Underwriter', 'avatar': 'AK'},
    {'username': 'pholland', 'fullName': 'Patrick Holland', 'branch': 'CAROLINA', 'role': 'Underwriter', 'avatar': 'PH'},
    {'username': 'dbossen', 'fullName': 'Doug Bossen', 'branch': 'TWIN CITIES', 'role': 'Underwriter', 'avatar': 'DB'},
]

TARGET_BRANCHES_RAW = ['KENTUCKY', 'CINCINNATI', 'ORANGE', 'TENNESSEE', 'CHICAGO CONTRACT', 'CAROLINA', 'TWIN CITIES', 'FLORIDA', 'WALNUT CREEK']
TARGET_USERNAMES = [u['username'] for u in USER_PROFILES]

def strip_branch_code(branch_str):
    """Convert '1855 - CINCINNATI' to 'CINCINNATI'"""
    if not branch_str:
        return ''
    m = re.match(r'^\d+\s*-\s*(.+)$', str(branch_str).strip())
    return m.group(1).strip() if m else str(branch_str).strip()

def normalize_branch(name):
    """Normalize branch name to uppercase"""
    if not name:
        return ''
    return str(name).strip().upper()

def fmt_date(val):
    """Convert various date formats to MM/DD/YYYY string"""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%m/%d/%Y')
    s = str(val).strip()
    if not s or s.lower() in ('none', 'n/a', 'null'):
        return ''
    return s

def parse_dollar(val):
    """Parse dollar amount from string or number"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace('$', '').replace(',', '').strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0

def safe_str(val):
    """Safely convert to string"""
    if val is None:
        return ''
    return str(val).strip()

def safe_float(val):
    """Safely convert to float"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(',', '').replace('%', '').strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None

# ==================== EXTRACT: Account Review Rating Report ====================
def extract_arr():
    print("Extracting Account Review Rating Report...")
    wb = openpyxl.load_workbook(f'{DATA_DIR}/Account Review Rating Report.xlsx', read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    results = []
    for row in rows:
        if len(row) < 11:
            continue
        branch_raw, account, acct_num, fs_type, fs_date, review_type, review_state, arr_created, reviewer, arr_approval, grade = row[:11]
        branch = strip_branch_code(safe_str(branch_raw))
        if not account or not branch:
            continue
        results.append({
            'branch': branch,
            'account': safe_str(account),
            'accountNumber': safe_str(acct_num),
            'fsType': safe_str(fs_type),
            'fsDate': fmt_date(fs_date),
            'type': safe_str(review_type),
            'reviewState': safe_str(review_state),
            'arrCreatedDate': fmt_date(arr_created),
            'assignee': safe_str(reviewer),
            'arrApprovalDate': fmt_date(arr_approval),
            'grade': safe_str(grade) if grade else '',
        })
    print(f"  -> {len(results)} ARR records extracted")
    return results

# ==================== EXTRACT: BondBox All Bonds ====================
def extract_bonds():
    print("Extracting BondBox All Bonds (filtered by branch)...")
    wb = openpyxl.load_workbook(f'{DATA_DIR}/BondBox All Bonds.xlsx', read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    seen = {}  # bond_number -> record (keep latest by created date)
    for row in rows:
        if len(row) < 10:
            continue
        bond_num, created, effective, expiration, agency_code, agency, branch_code, branch_name, underwriter, bond_class = row[:10]
        branch = normalize_branch(branch_name)
        if branch not in TARGET_BRANCHES_RAW:
            continue
        bn = safe_str(bond_num)
        if not bn:
            continue
        rec = {
            'bondNumber': bn,
            'dateCreated': fmt_date(created),
            'effectiveDate': fmt_date(effective),
            'expirationDate': fmt_date(expiration),
            'agencyCode': safe_str(agency_code),
            'agency': safe_str(agency),
            'branchCode': safe_str(branch_code),
            'branch': branch,
            'underwriter': safe_str(underwriter),
            'bondClass': safe_str(bond_class),
        }
        # Keep latest version per bond number
        if bn not in seen:
            seen[bn] = rec
        else:
            existing_date = seen[bn].get('dateCreated', '')
            new_date = rec.get('dateCreated', '')
            if new_date > existing_date:
                seen[bn] = rec

    results = list(seen.values())
    print(f"  -> {len(results)} unique bonds extracted (from {len(rows)} rows)")
    return results

# ==================== EXTRACT: Line of Authority Report ====================
def extract_loa():
    print("Extracting Line of Authority Report...")
    wb = openpyxl.load_workbook(f'{DATA_DIR}/Line of Authority Report.xlsx', read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the header row (contains 'Branch' or 'Account Name')
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(str(c).strip() == 'Branch' for c in row if c):
            header_idx = i
            break

    results = []
    for row in rows[header_idx + 1:]:
        if len(row) < 9:
            continue
        branch, account, acct_num, cust_num, status, exp_date, aggregate, single, user = row[:9]
        if not account:
            continue
        results.append({
            'branch': normalize_branch(branch),
            'account': safe_str(account),
            'accountNumber': safe_str(acct_num),
            'customerNumber': safe_str(cust_num),
            'status': safe_str(status),
            'expDate': fmt_date(exp_date),
            'aggregate': parse_dollar(aggregate),
            'single': parse_dollar(single),
            'user': safe_str(user),
        })
    print(f"  -> {len(results)} LOA records extracted")
    return results

# ==================== EXTRACT: Red Flag Audit Report ====================
def extract_red_flags():
    print("Extracting Red Flag Audit Report...")
    wb = openpyxl.load_workbook(f'{DATA_DIR}/Red Flag Audit Report.xlsx', read_only=True, data_only=True)

    result = {}
    for sheet_name in wb.sheetnames:
        # Only process branch-specific sheets (format: 'NNNN - NAME')
        if not re.match(r'^\d+\s*-\s*', sheet_name):
            continue

        branch = strip_branch_code(sheet_name)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Find header row containing 'Account Name'
        header_idx = 0
        for i, row in enumerate(rows):
            if row and any(str(c).strip() == 'Account Name' for c in row if c):
                header_idx = i
                break

        accounts = []
        for row in rows[header_idx + 1:]:
            if len(row) < 20:
                continue
            # Columns: 0=Branch, 2=Account Name, 4=FS Date, 5=FYE, 7=Z Score, 8=Net Quick,
            # 9=Net Worth, 10=D/E, 11=UB/NW, 12=UB/NQ, 13=Net Cash, 14=Net Income,
            # 15=NQ/LOA, 16=NW/LOA, 17=NQ/WOH, 18=NW/WOH, 19=Count
            account = row[2] if len(row) > 2 else None
            if not account:
                continue
            accounts.append({
                'account': safe_str(account),
                'financialStatement': fmt_date(row[4] if len(row) > 4 else None),
                'fye': safe_str(row[5] if len(row) > 5 else None),
                'zScore': safe_float(row[7] if len(row) > 7 else None),
                'netQuick': safe_float(row[8] if len(row) > 8 else None),
                'netWorth': safe_float(row[9] if len(row) > 9 else None),
                'debtEquity': safe_float(row[10] if len(row) > 10 else None),
                'ubNW': safe_float(row[11] if len(row) > 11 else None),
                'ubNQ': safe_float(row[12] if len(row) > 12 else None),
                'netCash': safe_float(row[13] if len(row) > 13 else None),
                'netIncome': safe_float(row[14] if len(row) > 14 else None),
                'nqLOA': safe_float(row[15] if len(row) > 15 else None),
                'nwLOA': safe_float(row[16] if len(row) > 16 else None),
                'nqWOH': safe_float(row[17] if len(row) > 17 else None),
                'nwWOH': safe_float(row[18] if len(row) > 18 else None),
                'flagCount': int(row[19]) if len(row) > 19 and row[19] else 0,
            })
        if accounts:
            result[branch] = accounts

    wb.close()
    print(f"  -> {len(result)} branches, {sum(len(v) for v in result.values())} red flag accounts extracted")
    return result

# ==================== EXTRACT: Timeliness of Financial Statements ====================
def extract_timeliness():
    print("Extracting Timeliness of Financial Statements (Detail sheet)...")
    wb = openpyxl.load_workbook(f'{DATA_DIR}/Timeliness of Financial Statements Report.xlsx', read_only=True, data_only=True)

    # Read detail sheet
    ws = wb['Report Details'] if 'Report Details' in wb.sheetnames else wb.worksheets[1]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    results = []
    for row in rows:
        if len(row) < 7:
            continue
        branch, account, acct_num, underwriter, prof_status, ni_2025, ni_2024 = row[:7]
        if not account:
            continue
        results.append({
            'branch': normalize_branch(branch),
            'account': safe_str(account),
            'accountNumber': safe_str(acct_num),
            'underwriter': safe_str(underwriter),
            'profitabilityStatus': safe_str(prof_status),
            'netIncome2025': parse_dollar(ni_2025) if ni_2025 and str(ni_2025).strip().lower() != 'n/a' else None,
            'netIncome2024': parse_dollar(ni_2024) if ni_2024 and str(ni_2024).strip().lower() != 'n/a' else None,
        })
    print(f"  -> {len(results)} financial timeliness records extracted")
    return results

# ==================== EXTRACT: Active Agencies ====================
def extract_agencies():
    print("Extracting Active Agencies...")
    wb = openpyxl.load_workbook(f'{DATA_DIR}/Active Agencies and Tiered Commission Report.xlsx', read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the actual data header (row with 'Branch Name' or 'Agency Name')
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(str(c).strip() == 'Agency Name' for c in row if c):
            header_idx = i
            break

    results = []
    for row in rows[header_idx + 1:]:
        if len(row) < 8:
            continue
        branch, agency, code, comm_purch, comm_renew, contract_purch, contract_renew, status = row[:8]
        if not agency:
            continue
        results.append({
            'branch': normalize_branch(branch),
            'agency': safe_str(agency),
            'agencyCode': safe_str(code),
            'commercialPurchaseRate': safe_str(comm_purch),
            'commercialRenewalRate': safe_str(comm_renew),
            'contractPurchaseRate': safe_str(contract_purch),
            'contractRenewalRate': safe_str(contract_renew),
            'status': safe_str(status),
        })
    print(f"  -> {len(results)} agency records extracted")
    return results

# ==================== BUILD USERNAME -> FULL NAME MAP ====================
def build_username_map(timeliness_data, arr_data, loa_data):
    """Cross-reference timeliness (full names) with ARR/LOA (usernames) via branch+account overlap"""
    print("Building username -> full name map...")

    # Build branch->accounts from ARR data (keyed by username)
    user_branch_accounts = defaultdict(lambda: defaultdict(set))
    for r in arr_data:
        if r['assignee']:
            user_branch_accounts[r['assignee']][r['branch']].add(r['account'])

    # Build branch->accounts from LOA data (keyed by username)
    for r in loa_data:
        if r['user']:
            user_branch_accounts[r['user']][r['branch']].add(r['account'])

    # Build branch->accounts from timeliness (keyed by full name)
    fullname_branch_accounts = defaultdict(lambda: defaultdict(set))
    for r in timeliness_data:
        if r['underwriter']:
            fullname_branch_accounts[r['underwriter']][r['branch']].add(r['account'])

    # Match by branch+account overlap
    username_map = {}
    for profile in USER_PROFILES:
        username_map[profile['username']] = profile['fullName']

    # Also try to map additional usernames
    for uname, branch_accts in user_branch_accounts.items():
        if uname in username_map:
            continue
        best_match = None
        best_overlap = 0
        for fname, fn_branch_accts in fullname_branch_accounts.items():
            overlap = 0
            for branch in branch_accts:
                if branch in fn_branch_accts:
                    overlap += len(branch_accts[branch] & fn_branch_accts[branch])
            if overlap > best_overlap:
                best_overlap = overlap
                best_match = fname
        if best_match and best_overlap >= 2:
            username_map[uname] = best_match

    print(f"  -> {len(username_map)} username mappings built")
    return username_map

# ==================== GENERATE data.js ====================
def generate_js(arr_data, bonds_data, loa_data, red_flags, timeliness_data, agencies_data, username_map):
    print("Generating data.js...")

    lines = []
    lines.append('/* ========================================')
    lines.append('   BondBox Prototype - Real Data Layer')
    lines.append('   Auto-generated from Excel reports')
    lines.append(f'   Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    lines.append('   ======================================== */')
    lines.append('')

    # User profiles (small, keep readable)
    lines.append('// ==================== USER PROFILES ====================')
    lines.append(f'const USER_PROFILES = {json.dumps(USER_PROFILES, indent=2)};')
    lines.append('')

    # Username map (small, keep readable)
    lines.append('// ==================== USERNAME MAP ====================')
    lines.append(f'const USERNAME_MAP = {json.dumps(username_map, indent=2)};')
    lines.append('')

    # ARR data (compact for size)
    lines.append('// ==================== ACCOUNT REVIEW RATINGS ====================')
    lines.append(f'const realARRs = {json.dumps(arr_data, separators=(",", ":"))};')
    lines.append('')

    # Bonds data (compact)
    lines.append('// ==================== BONDS ====================')
    lines.append(f'const realBonds = {json.dumps(bonds_data, separators=(",", ":"))};')
    lines.append('')

    # LOA data (compact)
    lines.append('// ==================== LINES OF AUTHORITY ====================')
    lines.append(f'const realLOAData = {json.dumps(loa_data, separators=(",", ":"))};')
    lines.append('')

    # Red flags (compact)
    lines.append('// ==================== RED FLAG AUDIT ====================')
    lines.append(f'const realRedFlagData = {json.dumps(red_flags, separators=(",", ":"))};')
    lines.append('')

    # Financial timeliness (compact)
    lines.append('// ==================== FINANCIAL TIMELINESS ====================')
    lines.append(f'const realFinancialTimeliness = {json.dumps(timeliness_data, separators=(",", ":"))};')
    lines.append('')

    # Agencies (compact)
    lines.append('// ==================== ACTIVE AGENCIES ====================')
    lines.append(f'const realAgencies = {json.dumps(agencies_data, separators=(",", ":"))};')
    lines.append('')

    js_content = '\n'.join(lines)

    output_path = r'C:\Development\New Extension Build\data.js'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

    size_kb = len(js_content) / 1024
    print(f"  -> data.js written: {size_kb:.0f} KB, {len(lines)} lines")
    return output_path

# ==================== MAIN ====================
if __name__ == '__main__':
    print("=" * 60)
    print("BondBox Data Extraction")
    print("=" * 60)

    arr_data = extract_arr()
    bonds_data = extract_bonds()
    loa_data = extract_loa()
    red_flags = extract_red_flags()
    timeliness_data = extract_timeliness()
    agencies_data = extract_agencies()
    username_map = build_username_map(timeliness_data, arr_data, loa_data)

    output = generate_js(arr_data, bonds_data, loa_data, red_flags, timeliness_data, agencies_data, username_map)

    print("")
    print("=" * 60)
    print(f"Done! Output: {output}")
    print("=" * 60)
